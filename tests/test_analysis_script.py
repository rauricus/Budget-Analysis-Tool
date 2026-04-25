#!/usr/bin/env python3
"""Test category analysis script."""
import sys
import os
import tempfile
from pathlib import Path
import pandas as pd

# Add root directory to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from openpyxl import load_workbook
from analyze_by_category import (
    load_categorized_csv,
    load_dataset_categorized_csvs,
    analyze_by_category,
    analyze_by_subcategory,
    create_excel_report,
    load_months_metadata,
)


def test_load_categorized_csv():
    """Test that categorized CSV is loaded correctly."""
    csv_path = 'data/example/output/export.202503.categorized.csv'

    df = load_categorized_csv(csv_path)

    assert df is not None, "CSV should be loaded"
    assert len(df) > 0, "CSV should contain transactions"
    assert 'Category' in df.columns, "Should have Category column"
    assert 'Subcategory' in df.columns, "Should have Subcategory column"
    assert 'Credit in CHF' in df.columns, "Should have Credit in CHF column"
    assert 'Debit in CHF' in df.columns, "Should have Debit in CHF column"


def test_analyze_by_category():
    """Test category analysis."""
    csv_path = 'data/example/output/export.202503.categorized.csv'
    df = load_categorized_csv(csv_path)

    category_stats = analyze_by_category(df)

    assert category_stats is not None, "Should return category statistics"
    assert len(category_stats) > 0, "Should have at least one category"
    assert 'Category' in category_stats.columns, "Should have Category column"
    assert 'Credit in CHF' in category_stats.columns, "Should have Credit in CHF column"
    assert 'Debit in CHF' in category_stats.columns, "Should have Debit in CHF column"
    assert 'Net in CHF' in category_stats.columns, "Should have Net in CHF column"

    # Verify totals match original data
    total_credit = df['Credit in CHF'].sum()
    total_debit = df['Debit in CHF'].sum()

    assert abs(category_stats['Credit in CHF'].sum() - total_credit) < 0.01, "Credit totals should match"
    assert abs(category_stats['Debit in CHF'].sum() - total_debit) < 0.01, "Debit totals should match"


def test_analyze_by_subcategory():
    """Test subcategory analysis."""
    csv_path = 'data/example/output/export.202503.categorized.csv'
    df = load_categorized_csv(csv_path)

    subcategory_stats = analyze_by_subcategory(df)

    assert subcategory_stats is not None, "Should return subcategory statistics"
    assert 'Category' in subcategory_stats.columns, "Should have Category column"
    assert 'Subcategory' in subcategory_stats.columns, "Should have Subcategory column"
    assert 'Credit in CHF' in subcategory_stats.columns, "Should have Credit in CHF column"
    assert 'Debit in CHF' in subcategory_stats.columns, "Should have Debit in CHF column"
    assert 'Net in CHF' in subcategory_stats.columns, "Should have Net in CHF column"

    # Verify no empty subcategories
    assert all(subcategory_stats['Subcategory'] != ''), "Should not have empty subcategories"


def test_category_uncategorized_handling():
    """Test that uncategorized transactions are properly handled."""
    csv_path = 'data/example/output/export.202503.categorized.csv'
    df = load_categorized_csv(csv_path)

    # Check if there are any uncategorized transactions
    uncategorized = df[df['Category'] == 'Uncategorized']

    if len(uncategorized) > 0:
        category_stats = analyze_by_category(df)

        # Should have an 'Uncategorized' category in stats
        assert 'Uncategorized' in category_stats['Category'].values, \
            "Should include Uncategorized category if present"


def test_numerical_calculations():
    """Test that numerical calculations are correct."""
    csv_path = 'data/example/output/export.202503.categorized.csv'
    df = load_categorized_csv(csv_path)

    category_stats = analyze_by_category(df)

    # Verify Net = Credit - Debit for each category
    for _, row in category_stats.iterrows():
        expected_net = row['Credit in CHF'] - row['Debit in CHF']
        assert abs(row['Net in CHF'] - expected_net) < 0.01, \
            f"Net should equal Credit - Debit for {row['Category']}"


def test_excel_report_creation():
    """Integration test for Excel report generation."""
    csv_path = 'data/example/output/export.202503.categorized.csv'
    df = load_categorized_csv(csv_path)

    category_stats = analyze_by_category(df)

        # Create temporary directory and file
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / 'subdir' / 'test_analysis.xlsx'
        months = sorted(df['Date'].dt.strftime("%Y-%m").dropna().unique().tolist())

        # Test that directory creation works
        create_excel_report(
            df,
            category_stats,
            str(output_path),
            'test.csv',
            months,
        )

        # Verify file was created
        assert output_path.exists(), "Excel file should be created"

        # Load and verify workbook structure
        wb = load_workbook(output_path)

        # Check expected sheets exist
        assert 'Summary' in wb.sheetnames, "Should have Summary sheet"
        assert 'Overviews by category' in wb.sheetnames, "Should have Overviews by category sheet"
        assert 'Category Analysis' in wb.sheetnames, "Should have Category Analysis sheet"
        assert 'Subcategory Analysis' in wb.sheetnames, "Should have Subcategory Analysis sheet"

        # Verify Summary sheet structure
        ws_summary = wb['Summary']
        assert ws_summary['A1'].value == 'Budget Analysis Summary', "Should have summary title"
        assert ws_summary['A1'].font.size == 16, "Summary main title should use size 16"
        assert ws_summary['A3'].value is None and ws_summary['A4'].value is None, "Summary should keep two blank rows after source"
        assert ws_summary['A5'].value == 'Overall Summary', "Should have overall summary section"
        assert ws_summary['A5'].font.size == 14, "Overall summary subtitle should use size 14"
        assert ws_summary['A7'].value == 'Transaction Category', "Should have transaction category table header"
        assert ws_summary['B7'].value == 'Credit (CHF)', "Should have credit header"
        assert ws_summary['C7'].value == 'Debit (CHF)', "Should have debit header"
        assert ws_summary['A8'].value == 'Income', "Should list income row"
        assert ws_summary['A9'].value == 'Expense', "Should list expense row"
        assert ws_summary['A10'].value == 'Refund', "Should list refund row"
        assert ws_summary['A11'].value == 'Total', "Should have total row"
        assert ws_summary['A13'].value == 'Transfer', "Should show transfer row with spacing"
        assert ws_summary['A15'].value == 'Grand Total', "Should have grand total row"

        # Verify first overview sheet has expected headers
        ws_overview = wb['Overviews by category']
        assert ws_overview['A1'].value == 'Budget Analysis by Category', "Should have title"
        assert ws_overview['A1'].font.size == 16, "Overview main title should use size 16"
        assert ws_overview['A3'].value is None and ws_overview['A4'].value is None, "Overview should keep two blank rows after source"
        assert ws_overview['A5'].value == 'Income by Category', "Should have income section"
        assert ws_overview['A5'].font.size == 14, "Overview subtitle should use size 14"

        # Verify income table has proper headers
        # Income table starts at row 7 (A5 is section title, A6 is blank, A7 is header row)
        assert ws_overview['A7'].value == 'Category', "Should have Category header for income table"
        assert ws_overview['B7'].value == 'Amount (CHF)', "Should have Amount header for income table"

        # Verify fixed spacing: next table header is 22 rows below previous table header
        assert ws_overview['A27'].value == 'Expenses by Category', "Should have Expenses section title at row 27"
        assert ws_overview['A29'].value == 'Category', "Should have Category header for expense table"
        assert ws_overview['B29'].value == 'Amount (CHF)', "Should have Amount header for expense table"

        # Verify fixed spacing: refund section title and header positions
        assert ws_overview['A49'].value == 'Refunds by Category', "Should have Refunds section title at row 49"
        if ws_overview['A51'].value == 'No refund data available.':
            pass
        else:
            assert ws_overview['A51'].value == 'Category', "Should have Category header for refund table"
            assert ws_overview['B51'].value == 'Amount (CHF)', "Should have Amount header for refund table"

        # Verify Category Analysis sheet has data
        ws_category = wb['Category Analysis']
        assert ws_category['A1'].value == 'Category Analysis', "Should have title"
        assert ws_category['A1'].font.size == 16, "Category sheet main title should use size 16"
        assert ws_category['A2'].value == "Note: Transactions with Transaction Category 'transfer' are excluded.", \
            "Should show transfer exclusion note on category sheet"
        assert ws_category['A3'].value is None and ws_category['A4'].value is None, "Category sheet should keep two blank rows after note"
        # A5 is the month label, A7 is the table header row
        assert ws_category['A5'].font.size == 14, "Month subtitle should use size 14"
        assert ws_category['A7'].value == 'Category', "Should have Category header"

        ws_subcategory = wb['Subcategory Analysis']
        assert ws_subcategory['A1'].font.size == 16, "Subcategory sheet main title should use size 16"
        assert ws_subcategory['A2'].value == "Note: Transactions with Transaction Category 'transfer' are excluded.", \
            "Should show transfer exclusion note on subcategory sheet"
        assert ws_subcategory['A3'].value is None and ws_subcategory['A4'].value is None, "Subcategory sheet should keep two blank rows after note"

        wb.close()


def test_analysis_sheets_exclude_transfer_transactions():
    """Category and subcategory sheets should omit rows tagged as transfer."""
    df = pd.DataFrame([
        {
            'Date': pd.Timestamp('2025-03-01'),
            'Category': 'Food',
            'Subcategory': 'Groceries',
            'Credit in CHF': 0.0,
            'Debit in CHF': 100.0,
            'Transaction Category': 'expense',
        },
        {
            'Date': pd.Timestamp('2025-03-02'),
            'Category': 'Transfers',
            'Subcategory': 'Account move',
            'Credit in CHF': 500.0,
            'Debit in CHF': 0.0,
            'Transaction Category': 'transfer',
        },
    ])
    category_stats = analyze_by_category(df)

    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / 'test_transfer_exclusion.xlsx'
        create_excel_report(df, category_stats, str(output_path), 'test', ['2025-03'])

        wb = load_workbook(output_path)
        ws_category = wb['Category Analysis']
        ws_subcategory = wb['Subcategory Analysis']

        category_values = [
            ws_category[f'A{row}'].value
            for row in range(1, 40)
            if ws_category[f'A{row}'].value is not None
        ]
        subcategory_values = [
            ws_subcategory[f'A{row}'].value
            for row in range(1, 40)
            if ws_subcategory[f'A{row}'].value is not None
        ]

        assert 'Transfers' not in category_values, "Transfer category should be excluded from category sheet"
        assert 'Transfers' not in subcategory_values, "Transfer category should be excluded from subcategory sheet"
        assert 'Food' in category_values, "Non-transfer category should remain in category sheet"

        wb.close()


def test_load_dataset_categorized_csvs_merges_all_files_and_adds_source():
    """Should merge all categorized CSV files from dataset output and add Source File column."""
    run_dir = Path('data/example')

    df, file_count = load_dataset_categorized_csvs(run_dir)

    assert file_count >= 1, "Should discover at least one categorized file"
    assert len(df) > 0, "Merged dataframe should contain transactions"
    assert 'Source File' in df.columns, "Merged dataframe should contain Source File column"
    assert df['Source File'].nunique() == file_count, "Each discovered file should be represented"


def test_load_dataset_categorized_csvs_raises_when_no_files():
    """Should raise if output directory has no categorized CSV files."""
    with tempfile.TemporaryDirectory() as tmpdir:
        run_dir = Path(tmpdir) / 'dataset'
        output_dir = run_dir / 'output'
        output_dir.mkdir(parents=True, exist_ok=True)

        try:
            load_dataset_categorized_csvs(run_dir)
            assert False, "Expected FileNotFoundError when no categorized files exist"
        except FileNotFoundError as exc:
            assert 'No categorized CSV files found' in str(exc)
