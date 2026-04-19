#!/usr/bin/env python3
"""
Budget Analysis by Category

Analyzes categorized CSV files and generates an Excel report with:
- Summary tables by category and subcategory
- Pie charts for visual analysis
- Flexibility for users to modify and customize

Usage:
    python analyze_by_category.py <run_dir> [output_excel_file]

Example:
    python analyze_by_category.py reference
    python analyze_by_category.py data/reference
    python analyze_by_category.py private my_analysis.xlsx
"""

import calendar
import json
import sys
from pathlib import Path
import numbers
from typing import Optional, Sequence, Tuple
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


REQUIRED_COLUMNS = {
    'Category',
    'Subcategory',
    'Credit in CHF',
    'Debit in CHF',
}

OVERVIEW_TABLE_HEADER_GAP = 22

def _month_label(month_str: str) -> str:
    """Convert 'YYYY-MM' string to English month name, e.g. '2024-01' -> 'January 2024'."""
    year, month = map(int, month_str.split("-"))
    return f"{calendar.month_name[month]} {year}"


def _resolve_run_directory(arg: str) -> Path:
    """Resolve run directory from CLI argument.

    Supports either a direct path (for example data/reference) or shorthand
    folder names under data/ (for example reference -> data/reference).
    """
    direct = Path(arg)
    if direct.exists() and direct.is_dir():
        return direct

    under_data = Path('data') / arg
    if under_data.exists() and under_data.is_dir():
        return under_data

    raise FileNotFoundError(
        f"Run directory not found: '{arg}' (also checked '{under_data}')"
    )


def load_categorized_csv(csv_path: str) -> pd.DataFrame:
    """Load a categorized CSV file.

    Args:
        csv_path: Path to the categorized CSV file

    Returns:
        DataFrame with transaction data
    """
    df = pd.read_csv(csv_path, sep=';', decimal=',', encoding='utf-8')

    # Parse date column
    df['Date'] = pd.to_datetime(df['Date'], format='%d.%m.%Y', errors='coerce')

    # Convert amount columns to numeric
    df['Credit in CHF'] = pd.to_numeric(df['Credit in CHF'], errors='coerce').fillna(0)
    df['Debit in CHF'] = pd.to_numeric(df['Debit in CHF'], errors='coerce').fillna(0)

    # Fill empty categories with "Uncategorized"
    df['Category'] = df['Category'].fillna('?').replace('?', 'Uncategorized')
    df['Subcategory'] = df['Subcategory'].fillna('')

    return df


def load_dataset_categorized_csvs(run_dir: Path) -> Tuple[pd.DataFrame, int]:
    """Load and merge all categorized CSV files from a run directory.

    Args:
        run_dir: Dataset run directory (for example data/reference)

    Returns:
        A tuple of merged DataFrame and number of loaded files

    Raises:
        FileNotFoundError: If output directory or categorized files are missing
        ValueError: If a file is missing required columns
        RuntimeError: If a categorized CSV cannot be loaded
    """
    output_dir = run_dir / 'output'
    if not output_dir.exists() or not output_dir.is_dir():
        raise FileNotFoundError(f"Output directory not found: {output_dir}")

    categorized_files = sorted(output_dir.glob('*.categorized.csv'))
    if not categorized_files:
        raise FileNotFoundError(f"No categorized CSV files found in: {output_dir}")

    dataframes = []
    for csv_file in categorized_files:
        try:
            df = load_categorized_csv(str(csv_file))
        except Exception as e:
            raise RuntimeError(f"Failed to load '{csv_file.name}': {e}") from e

        missing_columns = REQUIRED_COLUMNS.difference(df.columns)
        if missing_columns:
            missing_list = ', '.join(sorted(missing_columns))
            raise ValueError(
                f"File '{csv_file.name}' is missing required columns: {missing_list}"
            )

        df['Source File'] = csv_file.name
        dataframes.append(df)

    merged = pd.concat(dataframes, ignore_index=True)
    return merged, len(categorized_files)


def load_months_metadata(run_dir: Path) -> list[str]:
    """Load the months metadata written by categorize_transactions.py.

    Args:
        run_dir: Dataset run directory (e.g. data/reference)

    Returns:
        Sorted list of month strings in 'YYYY-MM' format

    Raises:
        FileNotFoundError: If months.json does not exist
    """
    months_path = run_dir / 'metadata' / 'months.json'
    if not months_path.exists():
        raise FileNotFoundError(
            f"Months metadata not found: {months_path}. "
            "Run categorize_transactions.py first to generate it."
        )
    with open(months_path, encoding='utf-8') as f:
        return json.load(f)


def analyze_by_category(df: pd.DataFrame) -> pd.DataFrame:
    """Analyze transactions by category.

    Args:
        df: DataFrame with transaction data

    Returns:
        DataFrame with category-level aggregations
    """
    category_stats = df.groupby('Category').agg({
        'Credit in CHF': 'sum',
        'Debit in CHF': 'sum'
    }).reset_index()

    # Calculate net (income - expenses)
    category_stats['Net in CHF'] = category_stats['Credit in CHF'] - category_stats['Debit in CHF']

    # Sort by total amount (debit + credit)
    category_stats['Total'] = category_stats['Credit in CHF'] + category_stats['Debit in CHF']
    category_stats = category_stats.sort_values('Total', ascending=False).drop('Total', axis=1)

    # Round to 2 decimals
    category_stats = category_stats.round(2)

    return category_stats


def analyze_by_subcategory(df: pd.DataFrame) -> pd.DataFrame:
    """Analyze transactions by category and subcategory.

    Args:
        df: DataFrame with transaction data

    Returns:
        DataFrame with subcategory-level aggregations
    """
    # Filter out rows without subcategory
    df_with_sub = df[df['Subcategory'] != ''].copy()

    if df_with_sub.empty:
        return pd.DataFrame(columns=['Category', 'Subcategory', 'Credit in CHF', 'Debit in CHF', 'Net in CHF'])

    subcategory_stats = df_with_sub.groupby(['Category', 'Subcategory']).agg({
        'Credit in CHF': 'sum',
        'Debit in CHF': 'sum'
    }).reset_index()

    # Calculate net
    subcategory_stats['Net in CHF'] = subcategory_stats['Credit in CHF'] - subcategory_stats['Debit in CHF']

    # Sort by category and total amount
    subcategory_stats['Total'] = subcategory_stats['Credit in CHF'] + subcategory_stats['Debit in CHF']
    subcategory_stats = subcategory_stats.sort_values(['Category', 'Total'], ascending=[True, False]).drop('Total', axis=1)

    # Round to 2 decimals
    subcategory_stats = subcategory_stats.round(2)

    return subcategory_stats


def create_excel_report(df: pd.DataFrame, category_stats: pd.DataFrame,
                        output_path: str, source_label: str, months: list[str]):
    """Create an Excel report with tables and charts.

    Args:
        df: Full transaction DataFrame with parsed Date column
        category_stats: DataFrame with overall category aggregations (for Overview sheet)
        output_path: Path to save the Excel file
        source_label: Human-readable source label for report header
        months: Sorted list of 'YYYY-MM' strings for per-month breakdown
    """
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create summary sheet
    ws_summary = wb.create_sheet("Summary", 0)
    _create_summary_sheet(ws_summary, df, source_label)

    # Create first overview sheet
    ws_overview = wb.create_sheet("Overviews by category", 1)
    _create_overview_sheet(ws_overview, df, source_label)

    # Create Category Analysis sheet (one table per month)
    ws_category = wb.create_sheet("Category Analysis", 2)
    _create_category_sheet(ws_category, df, months)

    # Create Subcategory Analysis sheet (one table per month)
    ws_subcategory = wb.create_sheet("Subcategory Analysis", 3)
    _create_subcategory_sheet(ws_subcategory, df, months)

    # Ensure output directory exists before saving
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    # Save workbook
    wb.save(output_file)
    print(f"✓ Excel report saved to: {output_file}")


def _create_summary_sheet(ws, df: pd.DataFrame, source_label: str):
    """Create Summary sheet with transaction-category overview and chart."""
    ws['A1'] = 'Budget Analysis Summary'
    ws['A1'].font = Font(size=16, bold=True)

    ws['A2'] = f'Source: {source_label}'
    ws['A2'].font = Font(size=10, italic=True)

    ws['A4'] = 'Overall Summary'
    ws['A4'].font = Font(size=14, bold=True)

    header_row = 6
    headers = ['Transaction Category', 'Credit (CHF)', 'Debit (CHF)']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(color='FFFFFF', bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    if 'Transaction Category' in df.columns:
        grouped = (
            df.assign(_tc=df['Transaction Category'].fillna('').astype(str).str.lower())
            .groupby('_tc', as_index=False)
            .agg({'Credit in CHF': 'sum', 'Debit in CHF': 'sum'})
        )
    else:
        grouped = pd.DataFrame(columns=['_tc', 'Credit in CHF', 'Debit in CHF'])

    def sums_for(tc: str) -> tuple[float, float]:
        row = grouped[grouped['_tc'] == tc]
        if row.empty:
            return 0.0, 0.0
        return float(row['Credit in CHF'].iloc[0]), float(row['Debit in CHF'].iloc[0])

    income_credit, income_debit = sums_for('income')
    expense_credit, expense_debit = sums_for('expense')
    refund_credit, refund_debit = sums_for('refund')
    transfer_credit, transfer_debit = sums_for('transfer')

    summary_rows = [
        (7, 'income', income_credit, income_debit),
        (8, 'expense', expense_credit, expense_debit),
        (9, 'refund', refund_credit, refund_debit),
    ]

    for row_idx, label, credit, debit in summary_rows:
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=credit).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=3, value=debit).number_format = '#,##0.00'

    total_credit = income_credit + expense_credit + refund_credit
    total_debit = income_debit + expense_debit + refund_debit
    ws.cell(row=10, column=1, value='Total').font = Font(bold=True)
    ws.cell(row=10, column=2, value=total_credit).number_format = '#,##0.00'
    ws.cell(row=10, column=3, value=total_debit).number_format = '#,##0.00'
    ws.cell(row=10, column=2).font = Font(bold=True)
    ws.cell(row=10, column=3).font = Font(bold=True)

    ws.cell(row=12, column=1, value='transfer')
    ws.cell(row=12, column=2, value=transfer_credit).number_format = '#,##0.00'
    ws.cell(row=12, column=3, value=transfer_debit).number_format = '#,##0.00'

    grand_total_credit = total_credit + transfer_credit
    grand_total_debit = total_debit + transfer_debit
    ws.cell(row=14, column=1, value='Grand Total').font = Font(bold=True)
    ws.cell(row=14, column=2, value=grand_total_credit).number_format = '#,##0.00'
    ws.cell(row=14, column=3, value=grand_total_debit).number_format = '#,##0.00'
    ws.cell(row=14, column=2).font = Font(bold=True)
    ws.cell(row=14, column=3).font = Font(bold=True)

    # Helper block for contiguous chart data (all four transaction categories)
    ws.cell(row=6, column=8, value='Transaction Category')
    ws.cell(row=6, column=9, value='Amount (CHF)')
    chart_rows = [
        ('income', income_credit + income_debit),
        ('expense', expense_credit + expense_debit),
        ('refund', refund_credit + refund_debit),
        ('transfer', transfer_credit + transfer_debit),
    ]
    for idx, (label, amount) in enumerate(chart_rows, start=7):
        ws.cell(row=idx, column=8, value=label)
        ws.cell(row=idx, column=9, value=amount).number_format = '#,##0.00'

    chart = PieChart()
    chart.title = 'Transaction Categories Overview'
    chart.style = 10
    chart.height = 10
    chart.width = 14
    labels = Reference(ws, min_col=8, min_row=7, max_row=10)
    chart_data = Reference(ws, min_col=9, min_row=6, max_row=10)
    chart.add_data(chart_data, titles_from_data=True)
    chart.set_categories(labels)

    # Keep chart aligned with table header row.
    ws.add_chart(chart, 'E6')

    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15


def _create_overview_sheet(ws, df: pd.DataFrame, source_label: str):
    """Create the overview sheet with summary and pie charts."""
    # Title
    ws['A1'] = 'Budget Analysis by Category'
    ws['A1'].font = Font(size=16, bold=True)

    ws['A2'] = f'Source: {source_label}'
    ws['A2'].font = Font(size=10, italic=True)

    # Build all three overview blocks with the same aggregation approach.
    income_data = _build_transaction_category_overview(
        df,
        transaction_category='income',
        amount_column='Income in CHF',
        amount_formula='credit_minus_debit',
    )
    expense_data = _build_transaction_category_overview(
        df,
        transaction_category='expense',
        amount_column='Expense in CHF',
        amount_formula='debit_minus_credit',
    )
    refund_data = _build_transaction_category_overview(
        df,
        transaction_category='refund',
        amount_column='Refund in CHF',
        amount_formula='credit_minus_debit',
    )

    # Income section
    current_row = 6
    ws[f'A{current_row}'] = 'Income by Category'
    ws[f'A{current_row}'].font = Font(size=12, bold=True)

    current_row += 2
    _add_table_and_chart(
        ws,
        income_data,
        start_row=current_row,
        amount_column='Income in CHF',
        chart_title='Income by Category',
        empty_message='No income data available.',
    )
    income_header_row = current_row

    # Expense section (header is fixed OVERVIEW_TABLE_HEADER_GAP rows below prior table header)
    expense_header_row = income_header_row + OVERVIEW_TABLE_HEADER_GAP
    current_row = expense_header_row - 2
    ws[f'A{current_row}'] = 'Expenses by Category'
    ws[f'A{current_row}'].font = Font(size=12, bold=True)

    current_row += 2
    _add_table_and_chart(
        ws,
        expense_data,
        start_row=current_row,
        amount_column='Expense in CHF',
        chart_title='Expenses by Category',
        empty_message='No expense data available.',
    )
    expense_header_row = current_row

    # Refund section (header is fixed OVERVIEW_TABLE_HEADER_GAP rows below prior table header)
    current_row = (expense_header_row + OVERVIEW_TABLE_HEADER_GAP) - 2
    ws[f'A{current_row}'] = 'Refunds by Category'
    ws[f'A{current_row}'].font = Font(size=12, bold=True)

    current_row += 2
    _add_table_and_chart(
        ws,
        refund_data,
        start_row=current_row,
        amount_column='Refund in CHF',
        chart_title='Refunds by Category',
        empty_message='No refund data available.',
    )

    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15


def _add_table_and_chart(
    ws,
    data: pd.DataFrame,
    start_row: int,
    amount_column: str,
    chart_title: str,
    empty_message: str,
):
    """Add category/amount table with blue header and pie chart next to it."""
    if data.empty:
        ws.cell(row=start_row, column=1, value=empty_message)
        return

    headers = ['Category', 'Amount (CHF)']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col)
        cell.value = header
        cell.font = Font(color='FFFFFF', bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

    for idx, (_, row) in enumerate(data.iterrows(), start=1):
        ws.cell(row=start_row + idx, column=1, value=row['Category'])
        ws.cell(row=start_row + idx, column=2, value=row[amount_column]).number_format = '#,##0.00'

    chart = PieChart()
    chart.title = chart_title
    chart.style = 10
    chart.height = 10
    chart.width = 14

    data_rows = len(data)
    labels = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + data_rows)
    chart_data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + data_rows)

    chart.add_data(chart_data, titles_from_data=True)
    chart.set_categories(labels)

    chart_cell = f"D{start_row}"
    ws.add_chart(chart, chart_cell)


def _build_transaction_category_overview(
    df: pd.DataFrame,
    transaction_category: str,
    amount_column: str,
    amount_formula: str,
) -> pd.DataFrame:
    """Aggregate one transaction category per category for the overview sheet."""
    if 'Transaction Category' not in df.columns:
        return pd.DataFrame(columns=['Category', amount_column])

    rows = df[df['Transaction Category'].fillna('').astype(str).str.lower() == transaction_category].copy()
    if rows.empty:
        return pd.DataFrame(columns=['Category', amount_column])

    grouped = rows.groupby('Category').agg({
        'Credit in CHF': 'sum',
        'Debit in CHF': 'sum',
    }).reset_index()

    if amount_formula == 'debit_minus_credit':
        grouped[amount_column] = grouped['Debit in CHF'] - grouped['Credit in CHF']
    else:
        grouped[amount_column] = grouped['Credit in CHF'] - grouped['Debit in CHF']

    grouped = grouped[grouped[amount_column] > 0].copy()
    grouped = grouped.sort_values(amount_column, ascending=False)
    return grouped[['Category', amount_column]]


def _exclude_transfer_transactions(df: pd.DataFrame) -> pd.DataFrame:
    """Return dataframe without rows classified as transaction category 'transfer'."""
    if 'Transaction Category' not in df.columns:
        return df
    mask = df['Transaction Category'].fillna('').astype(str).str.lower() != 'transfer'
    return df[mask].copy()


def _create_category_sheet(ws, df: pd.DataFrame, months: list[str]):
    """Create category analysis sheet with one table per month."""
    ws['A1'] = 'Category Analysis'
    ws['A1'].font = Font(size=14, bold=True)

    ws['A2'] = "Note: Transactions with Transaction Category 'transfer' are excluded."
    ws['A2'].font = Font(size=10, italic=True)

    analysis_df = _exclude_transfer_transactions(df)

    current_row = 3
    for month_str in months:
        year, month = map(int, month_str.split("-"))
        mask = (analysis_df['Date'].dt.year == year) & (analysis_df['Date'].dt.month == month)
        month_stats = analyze_by_category(analysis_df[mask])

        # Month header
        ws.cell(row=current_row, column=1, value=_month_label(month_str)).font = Font(size=12, bold=True)
        current_row += 2

        # Table header + data rows
        for i, row in enumerate(dataframe_to_rows(month_stats, index=False, header=True)):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                if i == 0:  # header row
                    cell.font = Font(color='FFFFFF', bold=True)
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center')
                elif col_idx > 1 and isinstance(value, numbers.Number):
                    cell.number_format = '#,##0.00'
            current_row += 1

        current_row += 2  # spacing between months

    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15


def _create_subcategory_sheet(ws, df: pd.DataFrame, months: list[str]):
    """Create subcategory analysis sheet with one table per month."""
    ws['A1'] = 'Subcategory Analysis'
    ws['A1'].font = Font(size=14, bold=True)

    ws['A2'] = "Note: Transactions with Transaction Category 'transfer' are excluded."
    ws['A2'].font = Font(size=10, italic=True)

    analysis_df = _exclude_transfer_transactions(df)

    current_row = 3
    for month_str in months:
        year, month = map(int, month_str.split("-"))
        mask = (analysis_df['Date'].dt.year == year) & (analysis_df['Date'].dt.month == month)
        month_stats = analyze_by_subcategory(analysis_df[mask])

        if month_stats.empty:
            continue

        # Month header
        ws.cell(row=current_row, column=1, value=_month_label(month_str)).font = Font(size=12, bold=True)
        current_row += 2

        # Table header + data rows
        for i, row in enumerate(dataframe_to_rows(month_stats, index=False, header=True)):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                if i == 0:  # header row
                    cell.font = Font(color='FFFFFF', bold=True)
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center')
                elif col_idx > 2 and isinstance(value, numbers.Number):
                    cell.number_format = '#,##0.00'
            current_row += 1

        current_row += 2  # spacing between months

    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15


def main(argv: Optional[Sequence[str]] = None):
    """Main entry point."""
    argv = argv if argv is not None else sys.argv[1:]

    if len(argv) < 1 or len(argv) > 2:
        print("Usage: python analyze_by_category.py <run_dir> [output_excel_file]")
        print("\nExample:")
        print("  python analyze_by_category.py reference")
        print("  python analyze_by_category.py data/reference")
        print("  python analyze_by_category.py private my_analysis.xlsx")
        return 2

    try:
        run_dir = _resolve_run_directory(argv[0])
    except FileNotFoundError as e:
        print(f"❌ {e}")
        return 1

    # Determine output path
    if len(argv) == 2:
        output_path = argv[1]
    else:
        # Default: in dataset output directory
        output_path = str(run_dir / 'output' / 'dataset.analysis.xlsx')

    print("=" * 60)
    print("  Budget Analysis - Category Report Generator")
    print("=" * 60)
    print(f"\nInput dataset:  {run_dir}")
    print(f"Output: {output_path}")

    # Discover and load all categorized files in dataset output
    print("\n1. Discovering and loading categorized CSV files...")
    try:
        df, file_count = load_dataset_categorized_csvs(run_dir)
        print(f"   Loaded {len(df)} transactions from {file_count} file(s)")
    except Exception as e:
        print(f"❌ Error loading categorized files: {e}")
        return 1

    print("\n2. Analyzing by category...")
    category_stats = analyze_by_category(df)
    print(f"   Found {len(category_stats)} categories")

    print("\n3. Loading month metadata...")
    try:
        months = load_months_metadata(run_dir)
        print(f"   Found {len(months)} month(s): {', '.join(_month_label(m) for m in months)}")
    except FileNotFoundError as e:
        print(f"\u274c {e}")
        return 1

    print("\n4. Generating Excel report...")
    try:
        source_label = f"{run_dir} ({file_count} categorized file(s))"
        create_excel_report(
            df,
            category_stats,
            output_path,
            source_label,
            months,
        )
    except Exception as e:
        print(f"❌ Error creating Excel report: {e}")
        return 1

    print("\n" + "=" * 60)
    print("Analysis completed successfully!")
    print("=" * 60)

    return 0


if __name__ == "__main__":
    sys.exit(main())
